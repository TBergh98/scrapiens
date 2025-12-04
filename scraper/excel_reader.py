"""Excel reader module for loading website configurations."""

import re
from pathlib import Path
from typing import List, Dict, Tuple, Any
from urllib.parse import urlparse
import openpyxl
from utils.logger import get_logger
from config.settings import get_config

logger = get_logger(__name__)


def sanitize_domain_name(url: str) -> str:
    """
    Extract and sanitize domain name from URL for use as filename.
    
    Args:
        url: Full URL string
        
    Returns:
        Sanitized domain name
        
    Example:
        >>> sanitize_domain_name('https://example.com/path')
        'example_com'
    """
    parsed = urlparse(url)
    name = parsed.netloc if parsed.netloc else url
    # Replace non-alphanumeric characters with underscore
    name = re.sub(r'[^0-9a-zA-Z-_]', '_', name)
    return name


def read_sites_from_xlsx(
    xlsx_path: Path,
    row_range: Tuple[int, int],
    sheet_index: int = 1,
    url_column: int = 1
) -> List[Dict[str, Any]]:
    """
    Read website URLs from an Excel file.
    
    Args:
        xlsx_path: Path to Excel file
        row_range: Tuple of (start_row, end_row) - 1-based indexing for end, inclusive
        sheet_index: Index of sheet to read (0-based)
        url_column: Column number containing URLs (1-based)
        
    Returns:
        List of site configuration dictionaries with keys:
        - name: Sanitized domain name
        - url: Full URL
        - js: Boolean indicating if JavaScript rendering needed
        - next_selector: CSS selector for pagination button (None if no pagination)
        - max_pages: Maximum pages to scrape
        
    Raises:
        FileNotFoundError: If Excel file not found
        ValueError: If workbook doesn't have enough sheets
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    
    logger.info(f"Reading sites from {xlsx_path}, sheet {sheet_index}, rows {row_range[0]}-{row_range[1]}")
    
    sites = []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    
    if len(wb.sheetnames) <= sheet_index:
        raise ValueError(
            f"Workbook has only {len(wb.sheetnames)} sheet(s). "
            f"Cannot access sheet at index {sheet_index}."
        )
    
    ws = wb[wb.sheetnames[sheet_index]]
    
    # Convert to 1-based row numbers for openpyxl (add 1 to start, add 1 to end for inclusive range)
    start_row = row_range[0] + 1
    end_row = row_range[1] + 1
    
    for row_idx in range(start_row, end_row + 1):
        cell_value = ws.cell(row=row_idx, column=url_column).value
        
        if cell_value:
            url = str(cell_value).strip()
            
            if url:
                name = sanitize_domain_name(url)
                
                sites.append({
                    'name': name,
                    'url': url,
                    'js': False,  # Default: no JavaScript rendering
                    'next_selector': None,  # Default: no pagination
                    'max_pages': 1  # Default: single page only
                })
                
                logger.debug(f"Loaded site: {name} -> {url}")
    
    wb.close()
    logger.info(f"Loaded {len(sites)} sites from Excel")
    
    return sites


def load_sites_from_config(category: str = 'standard') -> List[Dict[str, Any]]:
    """
    Load sites from Excel using configuration settings.
    
    Args:
        category: Category of sites to load ('standard' or 'problematic')
        
    Returns:
        List of site configuration dictionaries
    """
    config = get_config()
    
    # Get Excel configuration
    excel_path = config.get_full_path('paths.excel_file')
    sheet_index = config.get('excel.sheet_index', 1)
    url_column = config.get('excel.url_column', 1)
    
    # Get row range for category
    row_range_config = config.get(f'excel.row_ranges.{category}')
    if not row_range_config:
        raise ValueError(f"Unknown category: {category}. Use 'standard' or 'problematic'.")
    
    row_range = tuple(row_range_config)
    
    logger.info(f"Loading {category} sites from configuration")
    
    return read_sites_from_xlsx(
        xlsx_path=excel_path,
        row_range=row_range,
        sheet_index=sheet_index,
        url_column=url_column
    )
